from __future__ import annotations

import csv
import pathlib
from typing import Any

from openpyxl import load_workbook


def read_tabular(path: str) -> tuple[list[str], list[dict[str, str]]]:
    p = pathlib.Path(path).resolve()
    if not p.exists():
        raise FileNotFoundError(f"File non trovato: {p}")

    ext = p.suffix.lower()
    if ext in [".csv", ".txt"]:
        return _read_csv(p)
    if ext in [".xlsx"]:
        return _read_xlsx(p)
    raise ValueError("Formato non supportato. Usa CSV o XLSX.")


def _read_csv(p: pathlib.Path) -> tuple[list[str], list[dict[str, str]]]:
    with p.open("r", encoding="utf-8", errors="replace", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\\t,")
        reader = csv.DictReader(f, dialect=dialect)
        headers = reader.fieldnames or []
        rows: list[dict[str, str]] = []
        for r in reader:
            rows.append({k: str(v or "").strip() for k, v in r.items() if k is not None})
    return headers, rows


def _read_xlsx(p: pathlib.Path) -> tuple[list[str], list[dict[str, str]]]:
    wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        return [], []
    headers = [str(h or "").strip() for h in header_row]
    headers = [h if h else f"COL_{i+1}" for i, h in enumerate(headers)]

    rows: list[dict[str, str]] = []
    for row in it:
        d: dict[str, str] = {}
        for i, h in enumerate(headers):
            v: Any = row[i] if i < len(row) else ""
            d[h] = str(v if v is not None else "").strip()
        # scarta righe vuote
        if any(v for v in d.values()):
            rows.append(d)
    return headers, rows

